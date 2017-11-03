# Pick n user
from steem.post import Post
from steem.blog import Blog
from openpyxl import load_workbook, Workbook
from functools import reduce

def load_express(filename='accounts.xlsx'):
    wb = load_workbook(filename)
    ws = wb.active
    accounts = []
    for row in ws.iter_rows():
        vals = list(map(lambda x:x.value, row))
        accounts.append(vals[0])
    return accounts

def write_winner(filename, comments, express_winner, normal_winner, express, voters_weight):
    wb = Workbook()
    ws = wb.active
    ws.append(["account", "timestamp", "weight", "express", "normal", "idx", "comment", "reply"])
    express_idx = 0
    normal_idx = 0
    for c in comments:
        if(c['author'] not in voters_weight):
            continue
        if c['author'] in express:
            if c['author'] in express_winner:
                express_idx = express_idx + 1
                comment = "(익스프레스)%s번째 낙찰자이십니다 오픈채팅에서 카카오톡 ID를 이야기해주시면 보내드립니다 https://open.kakao.com/o/s4QAQsB" % express_idx
                ws.append([c['author'], c['created'], voters_weight[c['author']], "o", None, express_idx, c['body'], comment])
            else:
                ws.append([c['author'], c['created'], voters_weight[c['author']], "o", None, express_idx, c['body'], " "])
        elif c['author'] in normal_winner:
            normal_idx = normal_idx + 1
            comment = "%s번째 낙찰자이십니다 오픈채팅에서 카카오톡 ID를 이야기해주시면 보내드립니다 https://open.kakao.com/o/s4QAQsB" % normal_idx
            ws.append([c['author'], c['created'], voters_weight[c['author']], None, "o", normal_idx, c['body'], comment])
        else:
            ws.append([c['author'], c['created'], voters_weight[c['author']], None, None, None, c['body'], " "])
    filename = "".join([c for c in filename if c.isalpha() or c.isdigit() or c==' ']).rstrip() + ".xlsx"
    wb.save(filename)
    return filename

def try_parse_int(val, default):
    try:
        cv = int(val)
    except:
        cv = default
    return cv

def try_parse_float(val, default):
    try:
        cv = float(val)
    except:
        cv = default
    return cv

userId = 'virus707'
express_accounts = load_express()
pn = try_parse_int(input("몇번째 포스트의 당첨자를 구하시겠습니까? (default:1) : "), 1)
b = Blog(userId)
for i in range(0,pn):
    p = next(b)
pj = p.export()
sbd = float(pj['pending_payout_value']['amount'])
rid = pj['root_identifier']
voters = list(map(lambda x: x['voter'], pj['active_votes']))
voters_weight = {}
for x in pj['active_votes']:
    voters_weight[x['voter']] = x['weight']
print(pj['root_title'])
print("현재 글의 보팅금액은 %s SBD 입니다." % str(sbd))
unit = float(input("현재 글의 상품 단가(단위:SBD)를 입력해 주십시오. (총금액/단가 사용자 선정) (예:3.42) : "))
comment_vote = try_parse_float(input("예상 댓글 셀프 보팅 금액을 임력해 주십시오. (기본 0.0) : "), 0.0)
alpha = try_parse_float(input("후원 금액을 입력해 주십시오 (기본 0.0) : "), 0.0)
ratio = try_parse_float(input("보팅금액 비율을 입력해 주십시오 (기본 0.375) : "), 0.375)
prize_pool = (alpha + (sbd + comment_vote) * ratio)
amount = int(prize_pool/unit)
print("후원금 %s + (글 보상 %s + 댓글 보상 %s) * 보팅금액 비율 %s = %s SBD 입니다." % (str(alpha), str(sbd), str(comment_vote), str(ratio), str(prize_pool)))
print("%s명 선정을 진행합니다." % str(amount))
comments = list(map(lambda x:x.export(), list(p.get_replies())))
# candidates = list(map(lambda x:x['author'], filter(lambda x: '참여' in x['body'], comments)))
candidates = list(map(lambda x: x['author'], comments))
pick_express = []
pick_normal = []
for a in candidates:
    if amount <= 0:
        continue
    if(a in express_accounts):
        if a == userId:
            continue
        print(a)
        pick_express.append(a)
        amount = amount - 1
for a in candidates:
    if amount <= 0:
        continue
    if(a not in express_accounts):
        print(a)
        pick_normal.append(a)
        amount = amount - 1
print("랜딩 참여자 %s명, 일반 참여자 %s명 선정되었습니다." % (str(len(pick_express)), str(len(pick_normal))))
filename = write_winner(rid, comments, pick_express, pick_normal, express_accounts, voters_weight)
input("선정이 완료되었습니다. %s 파일을 확인 바랍니다." % filename)